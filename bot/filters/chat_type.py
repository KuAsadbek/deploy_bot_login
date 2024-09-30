import openpyxl
from aiogram.filters import Filter
from aiogram.filters import BaseFilter
from aiogram.types import Message,ContentType
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

from utils.db.class_db import SQLiteCRUD


CHANEL_ID = -1002020757864
TEACHER_MOD = 'main_app_teachermod'
PARENTS_MOD = 'main_app_parentmod'
DESCR = 'main_app_descriptionmod'
USERMOD = 'main_app_usermod'
CAT = 'main_app_categorymod'
BUT = 'main_app_buttonmod'
SAVE_DATA = 'main_app_save_user_data'

db = SQLiteCRUD('./db.sqlite3')

def is_uzbek_number(phone_number):
        return phone_number.startswith('+998') or phone_number.startswith('998')

def get_text_and_language(user_record, lg_index):
        main = db.read(DESCR,where_clause=f'title_id = 1')
        if user_record == 'start':
            return main[0][lg_index]
        else:
            lg = user_record[0][lg_index]
            n = 2 if lg == 'ru' else 1
            return main[0][n], lg

def generate_unique_code(cod):
        base_code = cod
        counter = 0
        
        # Разделяем буквы и числа один раз
        letter_part, number_part = base_code.split('-')
        number_part = int(number_part)

        # Один запрос для проверки наличия кода сразу во всех моделях
        def is_code_exists(code):
            parent_exists = db.read(USERMOD,where_clause=f'code = "{code}"')
            return bool(parent_exists)

        # Проверяем сразу первый код
        if not is_code_exists(base_code):
            return base_code

        # Если код существует, увеличиваем счётчик и генерируем новый код
        while True:
            counter += 1
            new_code = f"{letter_part}-{number_part + counter}"
            
            if not is_code_exists(new_code):
                return new_code

# Функция для стилизации ячеек с заголовками
def style_header_cells(sheet, columns):
    header_font = Font(bold=True, size=12, color="FFFFFF")  # Белый жирный шрифт
    fill = PatternFill(fill_type="solid", start_color="4F81BD", end_color="4F81BD")  # Синий фон
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))  # Тонкие границы
    
    for col in columns:
        for cell in sheet[col][:1]:  # Только первая строка (заголовки)
            cell.font = header_font
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")  # Центровка

# Функция для добавления границ и выравнивания
def style_body_cells(sheet, columns):
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col in columns:
        sheet.column_dimensions[col].width = 20  # Задаем ширину столбцов
        for cell in sheet[col][1:]:  # Все строки, кроме первой (заголовков)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")  # Выравнивание по левому краю

# Функция для раскрашивания ячеек в колонке 'G' в зависимости от значения
def style_column_g(sheet):
    green_fill = PatternFill(fill_type="solid", start_color="00FF00", end_color="00FF00")  # Зеленый цвет
    red_fill = PatternFill(fill_type="solid", start_color="FF0000", end_color="FF0000")  # Красный цвет
    
    for cell in sheet['G'][1:]:  # Пропускаем заголовок (начинаем с 1 строки)
        if cell.value == 1:
            cell.fill = green_fill
        elif cell.value == 0:
            cell.fill = red_fill

# Функция для создания листа и стилизации
def create_sheet(workbook, sheet_title, data, headers):
    sheet = workbook.create_sheet(sheet_title)
    sheet.append(headers)
    for row in data:
        sheet.append(row)
    columns_to_style = ["A", "B", "C", "D", "E", "F", "G", "H"]
    style_header_cells(sheet, columns_to_style)
    style_body_cells(sheet, columns_to_style)
    style_column_g(sheet)

# Основная функция для создания Excel файла с улучшенной стилизацией
def create_excel_with_data(Students, Teacher, Parents, file_name="output.xlsx"):
    try:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        headers = ["id", "Telegram ID", "Name", "School", "City", "Number", "Payment", "Language"]
        
        # Создаем листы, если данные есть
        if Students:
            create_sheet(workbook, "Students", Students, headers)
        if Teacher:
            create_sheet(workbook, "Teachers", Teacher, headers)
        if Parents:
            create_sheet(workbook, "Parents", Parents, headers)
        
        # Сохраняем файл только один раз в конце
        workbook.save(file_name)
        print(f"Файл '{file_name}' успешно создан.")
        return file_name
    
    except Exception as e:
        print(f"Ошибка при создании Excel файла: {e}")
        return None

class chat_type_filter(Filter):
    def __init__(self,chat_types:list[str]) -> None:
        self.chat_types = chat_types

    async def __call__(self,message:Message) -> bool:
        return message.chat.type in self.chat_types

class MediaFilter(BaseFilter):
    async def __call__(self, message: Message) -> bool:
        file = message.content_type in [ContentType.PHOTO, ContentType.DOCUMENT]
        return file if file else 'False'